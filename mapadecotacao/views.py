import os
import pdfplumber
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .forms import InsumoForm, UploadFileForm, OrcamentoForm
from .models import Insumo, Preco, Empresa, ArquivoAnexado, Mapa
from django.contrib import messages
from django.http import HttpResponse, HttpResponseServerError
from django.template import loader
from django.contrib.auth.models import User
from django.urls import reverse
import statistics
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.core.files.storage import FileSystemStorage
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from io import BytesIO
from django.core.files.base import ContentFile
import shutil
import zipfile
from django.core.exceptions import ValidationError
from django.core.files import File
from django.conf import settings
import tempfile
from django.utils import timezone
from django.contrib.auth import logout
from django.shortcuts import redirect


def handle_uploaded_file(uploaded_file):
    upload_dir = 'uploads/'
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir)
    file_name = os.path.join(upload_dir, uploaded_file.name)
    with open(file_name, 'wb+') as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)
    return file_name


@login_required
def cadastrar_orcamento(request):
    if request.method == 'POST':
        form = OrcamentoForm(request.POST, request.FILES)
        if form.is_valid():
            status_preco = request.POST.get('status_preco')
            empresa = request.user.empresa

            # Obtendo o arquivo PDF anexado (assumindo que é um único arquivo PDF para todos os insumos)
            documento = request.FILES.get('documento')  # Assume que é um único arquivo PDF

            if not documento:
                messages.error(request, "Por favor, anexe um arquivo PDF.")
                return redirect('cadastrar_orcamento')

            # Itera sobre os insumos enviados
            for i in range(len(request.POST.getlist('codigo_insumo'))):
                codigo_insumo = request.POST.getlist('codigo_insumo')[i]
                nome_insumo = request.POST.getlist('nome_insumo')[i]
                unidade_medida = request.POST.getlist('unidade_medida')[i]
                preco = request.POST.getlist('preco')[i]

                if not codigo_insumo:
                    continue

                try:
                    # Obtém ou cria o insumo
                    insumo = Insumo.objects.get(codigo=codigo_insumo, empresa=empresa)
                except Insumo.DoesNotExist:
                    insumo = Insumo.objects.create(
                        codigo=codigo_insumo,
                        nome=nome_insumo,
                        unidade_medida=unidade_medida,
                        empresa=empresa
                    )

                # Criação do objeto Preco para o insumo
                preco_obj = Preco.objects.create(
                    insumo=insumo,
                    preco=preco,
                    cnpj=request.POST['cnpj'],
                    razao_social=request.POST['nome_empresa'],
                    vendedor=request.POST['vendedor'],
                    telefone=request.POST['telefone'],
                    data_cotacao=request.POST['data_cotacao'],
                    email=request.POST['email'],
                    status_preco=status_preco
                )

                try:
                    # Criar a pasta específica para o insumo (se não existir)
                    pasta_destino = os.path.join('arquivos_anexados', f"{insumo.codigo}-{insumo.nome}")
                    if not os.path.exists(pasta_destino):
                        os.makedirs(pasta_destino)  # Criação da pasta se não existir

                    # Caminho completo do arquivo de destino
                    pdf_dest_path = os.path.join(pasta_destino, documento.name)

                    # Movendo o PDF para a pasta correta
                    with open(pdf_dest_path, 'wb') as f:
                        for chunk in documento.chunks():
                            f.write(chunk)

                    # Agora, associe o arquivo à tabela ArquivoAnexado
                    ArquivoAnexado.objects.create(
                        empresa=empresa,
                        arquivo=documento,
                        preco=preco_obj  # Associando o arquivo ao preço
                    )

                except Exception as e:
                    messages.error(request, f"Erro ao salvar o arquivo: {e}")
                    return redirect('cadastrar_orcamento')

            messages.success(request, "Orçamento cadastrado com sucesso.")
            return redirect('cadastrar_orcamento')

        else:
            messages.error(request, "Formulário inválido. Verifique os dados enviados.")
    else:
        form = OrcamentoForm()

    return render(request, 'cadastrar_orcamento.html', {'form': form})

@login_required
def gerar_relatorio(request):
    empresa_do_usuario = request.user.empresa
    if empresa_do_usuario:
        if request.method == 'POST':
            calcular = request.POST.get('calcular')
            insumos_selecionados = request.POST.getlist('insumos')
            insumos_do_usuario = Insumo.objects.filter(id__in=insumos_selecionados)

            # Cria um novo workbook para o Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Relatório de Cotações'

            headers = [
                'Código', 'Nome', 'Unidade de Medida', 'Cálculo',
                'Qtd Cotações', 'Fornecedor', 'CNPJ', 'Contato', 'Valor'
            ]

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')

            # Cabeçalho
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            row_num = 2
            # Inicializa o buffer para o arquivo ZIP
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                for insumo in insumos_do_usuario:
                    # Obter as 3 cotações mais recentes
                    precos_recentes = Preco.objects.filter(insumo=insumo).order_by('-data_cotacao')[:3]

                    # Obter os IDs das 3 cotações mais recentes
                    precos_recentes_ids = precos_recentes.values_list('id', flat=True)

                    # Obter os PDFs associados às 3 cotações mais recentes
                    arquivos_recentes = ArquivoAnexado.objects.filter(preco_id__in=precos_recentes_ids)

                    preços = [preco.preco for preco in precos_recentes]
                    calculo = 0
                    if calcular == 'media' and preços:
                        calculo = statistics.mean(preços)
                    elif calcular == 'mediana' and preços:
                        calculo = statistics.median(preços)

                    ws.cell(row=row_num, column=1, value=insumo.codigo).border = thin_border
                    ws.cell(row=row_num, column=2, value=insumo.nome).border = thin_border
                    ws.cell(row=row_num, column=3, value=insumo.unidade_medida).border = thin_border
                    ws.cell(row=row_num, column=4, value=calcular).border = thin_border
                    ws.cell(row=row_num, column=5, value=len(precos_recentes)).border = thin_border

                    # Adicionar as cotações ao Excel
                    pdf_paths = []
                    for idx, preco in enumerate(precos_recentes):
                        ws.cell(row=row_num + idx, column=6, value=preco.razao_social).border = thin_border
                        ws.cell(row=row_num + idx, column=7, value=preco.cnpj).border = thin_border
                        ws.cell(row=row_num + idx, column=8, value=preco.telefone).border = thin_border
                        ws.cell(row=row_num + idx, column=9, value=preco.preco).border = thin_border

                        # Adicionar o PDF ao ZIP apenas se existir e for um dos 3 mais recentes
                        try:
                            arquivo_anexado = arquivos_recentes.get(preco=preco)
                            pdf_path = arquivo_anexado.arquivo.path
                            if os.path.exists(pdf_path):
                                pdf_paths.append(pdf_path)
                            else:
                                ws.cell(row=row_num + idx, column=10,
                                        value=f'Erro: Arquivo não encontrado').border = thin_border
                        except ArquivoAnexado.DoesNotExist:
                            ws.cell(row=row_num + idx, column=10,
                                    value=f'Erro: Arquivo não anexado').border = thin_border

                    # Adicionar apenas os 3 PDFs mais recentes ao ZIP
                    pdf_paths_sorted = sorted(pdf_paths, key=lambda p: os.path.getmtime(p), reverse=True)[:3]
                    for pdf_path in pdf_paths_sorted:
                        zip_file.write(pdf_path, os.path.basename(pdf_path))

                    row_num += max(1, len(precos_recentes)) + 1

                # Definir a largura das colunas
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # letra da coluna
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

            # Adicionar o Excel ao ZIP
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            zip_buffer.seek(0)
            zip_buffer.write(excel_buffer.read())

            # Preparar a resposta HTTP com o arquivo ZIP
            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename=relatorio_cotacoes.zip'
            return response

        else:
            insumos_do_usuario = Insumo.objects.filter(empresa=empresa_do_usuario)
            context = {'insumos_do_usuario': insumos_do_usuario}
            return render(request, 'relatorios.html', context)
    else:
        return render(request, 'error.html', {
            'error_message': 'Você não possui uma empresa associada. Por favor, entre em contato com o administrador.'
        })

def cadastrar_insumos(request):
    if request.method == 'POST':
        form = InsumoForm(request.POST, user=request.user)
        if form.is_valid():
            nome_insumo = form.cleaned_data['nome']
            insumo = form.save(commit=False)
            insumo.empresa = request.user.empresa
            insumo.user = request.user
            insumo.aprovado = 1 if request.user.is_staff else 0
            insumo.nome = nome_insumo
            insumo.save()
            success_message = f'O insumo "{nome_insumo}" foi cadastrado com sucesso!' if request.user.is_staff else f'O insumo "{nome_insumo}" foi enviado para avaliação.'
            messages.success(request, success_message)
            return redirect('cadastrar_insumos')
    else:
        form = InsumoForm(user=request.user)
    return render(request, "cadastrar_insumos.html", {"form": form})

@login_required
def delete_insumo(request, insumo_id):
    # Obter o objeto de Insumo ou retornar um 404 se não for encontrado
    insumo = get_object_or_404(Insumo, pk=insumo_id)

    # Lógica para excluir o insumo aqui
    insumo.delete()

    # Redirecionar para uma página de sucesso ou para onde for apropriado
    return redirect('user_profile')

@login_required
def avaliar_insumo(request, insumo_id):
    if request.user.is_staff:
        insumo = Insumo.objects.get(pk=insumo_id)
        if request.method == 'POST':
            aprovado = request.POST.get('aprovado')
            if aprovado == 'sim':
                insumo.aprovado = 1
                insumo.save()
            return redirect('insumos_para_avaliar')
        else:
            return render(request, 'avaliar_insumo.html', {'insumo': insumo})
    else:
        return redirect('user_profile')

@login_required
def insumos_para_avaliar(request):
    if request.user.is_staff and request.user.empresa:
        insumos_para_avaliar = Insumo.objects.filter(empresa=request.user.empresa, aprovado=0)
        return render(request, 'insumos_para_avaliar.html', {'insumos_para_avaliar': insumos_para_avaliar})
    else:
        return redirect('user_profile')


@login_required
def gerar_relatorio(request):
    empresa_do_usuario = request.user.empresa
    if empresa_do_usuario:
        if request.method == 'POST':
            calcular = request.POST.get('calcular')
            insumos_selecionados = request.POST.getlist('insumos')
            insumos_do_usuario = Insumo.objects.filter(id__in=insumos_selecionados)

            # Cria um workbook para o relatório Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Relatório de Cotações'

            headers = [
                'Código', 'Nome', 'Unidade de Medida', 'Cálculo',
                'Qtd Cotações', 'Fornecedor', 'CNPJ', 'Contato', 'Valor'
            ]

            # Define bordas e estilos
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')

            # Preenche o cabeçalho do Excel
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            row_num = 2
            pdf_paths = []  # Lista para armazenar os caminhos dos PDFs a serem incluídos no zip

            for insumo in insumos_do_usuario:
                preços_recentes = Preco.objects.filter(insumo=insumo).order_by('-data_cotacao')[:3]
                preços = [preco.preco for preco in preços_recentes]
                calculo = 0
                if calcular == 'media' and preços:
                    calculo = statistics.mean(preços)
                elif calcular == 'mediana' and preços:
                    calculo = statistics.median(preços)

                ws.cell(row=row_num, column=1, value=insumo.codigo).border = thin_border
                ws.cell(row=row_num, column=2, value=insumo.nome).border = thin_border
                ws.cell(row=row_num, column=3, value=insumo.unidade_medida).border = thin_border
                ws.cell(row=row_num, column=4, value=calcular).border = thin_border
                ws.cell(row=row_num + 1, column=4, value=calculo).border = thin_border
                ws.cell(row=row_num, column=5, value=len(preços_recentes)).border = thin_border

                # Adiciona dados do preço e caminhos dos PDFs ao zip
                for idx, preco in enumerate(preços_recentes):
                    ws.cell(row=row_num + idx, column=6, value=preco.razao_social).border = thin_border
                    ws.cell(row=row_num + idx, column=7, value=preco.cnpj).border = thin_border
                    ws.cell(row=row_num + idx, column=8, value=preco.telefone).border = thin_border
                    ws.cell(row=row_num + idx, column=9, value=preco.preco).border = thin_border

                    try:
                        arquivo_anexado = ArquivoAnexado.objects.get(preco=preco)
                        pdf_path = arquivo_anexado.arquivo.path  # Caminho do arquivo PDF
                        if os.path.exists(pdf_path):
                            pdf_paths.append(pdf_path)
                        else:
                            # Se o arquivo PDF não for encontrado, pode-se adicionar uma mensagem de erro no Excel se desejar
                            pass
                    except ArquivoAnexado.DoesNotExist:
                        pass

                row_num += max(1, len(preços_recentes)) + 1

            # Define largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Cria um buffer para o arquivo Excel
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Cria um arquivo zip em memória
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                # Adiciona o arquivo Excel ao zip
                zip_file.writestr('relatorio_cotacoes.xlsx', excel_buffer.read())

                # Adiciona os PDFs ao zip
                for pdf_path in pdf_paths:
                    with open(pdf_path, 'rb') as pdf_file:
                        zip_file.writestr(os.path.basename(pdf_path), pdf_file.read())

            zip_buffer.seek(0)

            response = HttpResponse(zip_buffer, content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename=relatorio_cotacoes.zip'
            return response
        else:
            insumos_do_usuario = Insumo.objects.filter(empresa=empresa_do_usuario)
            context = {'insumos_do_usuario': insumos_do_usuario}
            return render(request, 'relatorios.html', context)
    else:
        return render(request, 'error.html', {
            'error_message': 'Você não possui uma empresa associada. Por favor, entre em contato com o administrador.'
        })

def user_list_by_group(request, group_name):
    users = User.objects.filter(group=group_name)
    context = {'users': users}
    return render(request, 'template.html', context)


@login_required
def rename_map_title(request, mapa_id):
    mapa = get_object_or_404(Mapa, pk=mapa_id)  # Obtém o objeto Mapa com o ID fornecido

    if request.method == 'POST':
        new_title = request.POST.get('new_title')  # Obtém o novo título do mapa do POST request
        mapa.title = new_title  # Atualiza o título do mapa
        mapa.save()  # Salva as alterações no banco de dados
        return redirect('user_profile')  # Redireciona de volta para a página de perfil

    # Se o método não for POST, retorne algo ou redirecione para a página inicial
    return redirect('user_profile')  # Redireciona para a página de perfil se não for um POST request


from django.contrib.auth import get_user_model
import logging

logger = logging.getLogger(__name__)


User = get_user_model()


@login_required
def profile(request):
    logger.debug("Usuário autenticado: %s", request.user)

    # Garantir que o request.user é uma instância do modelo CustomUser
    user = request.user
    logger.debug("Verificando se o usuário tem a empresa associada: %s", user)

    if hasattr(user, 'empresa'):
        empresa = user.empresa
        logger.debug("Empresa associada ao usuário: %s", empresa)

        if empresa:
            # Processamento de POST para deletar insumo
            if request.method == 'POST':
                insumo_id_to_delete = request.POST.get('insumo_id_to_delete')
                logger.debug("Tentando deletar o insumo com ID: %s", insumo_id_to_delete)

                if insumo_id_to_delete:
                    try:
                        insumo_to_delete = Insumo.objects.get(id=insumo_id_to_delete)
                        logger.debug("Insumo encontrado: %s", insumo_to_delete)
                        insumo_to_delete.delete()
                        logger.debug("Insumo deletado com sucesso.")
                    except Insumo.DoesNotExist:
                        logger.error("Insumo não encontrado com o ID: %s", insumo_id_to_delete)
                    return redirect('profile')  # Redireciona de volta para a página de perfil

            # Obtendo os insumos associados à empresa
            insumos = Insumo.objects.filter(empresa=empresa)
            logger.debug("Insumos encontrados para a empresa %s: %s", empresa, insumos)

            insumos_com_preco = []
            for insumo in insumos:
                logger.debug("Buscando preços recentes para o insumo: %s", insumo.nome)
                preços_recentes = Preco.objects.filter(insumo=insumo).order_by('-data_cotacao')[:3]
                logger.debug("Preços recentes para o insumo '%s': %s", insumo.nome, preços_recentes)

                preços_e_razoes_sociais = []
                for preço in preços_recentes:
                    preços_e_razoes_sociais.append({
                        'preco': preço.preco,
                        'razao_social': preço.razao_social
                    })

                insumos_com_preco.append({
                    'insumo': insumo,
                    'precos_e_razoes_sociais': preços_e_razoes_sociais
                })

            # Obtendo os mapas da empresa
            mapas_da_empresa = Mapa.objects.filter(usuario=user, empresa=empresa)
            logger.debug("Mapas encontrados para a empresa %s: %s", empresa, mapas_da_empresa)

            # Contexto com as informações para renderizar a página de perfil
            context = {
                'insumos_com_preco': insumos_com_preco,
                'mapas_da_empresa': mapas_da_empresa
            }

            logger.debug("Renderizando página de perfil com o contexto: %s", context)
            return render(request, 'profile/profile.html', context)

    else:
        logger.error("Usuário não tem empresa associada.")
        return render(request, 'profile/error.html')

def aviso_insumos_vencidos(request):
    return render(request, 'aviso_insumos_vencidos.html')


def custom_logout(request):
    if request.method == 'POST':
        logout(request)
        return redirect('login')  # Redirecionar para a página de login após o logout
    else:
        # Se a solicitação não for POST, retorne um erro ou redirecione para a página inicial
        return redirect('/')  # Redirecionar para a página inicial



from zipfile import ZipFile

@login_required
def relatorios(request):
    empresa_do_usuario = request.user.empresa
    if empresa_do_usuario:
        if request.method == 'POST':
            calcular = request.POST.get('calcular')
            insumos_selecionados = request.POST.getlist('insumos')

            # Verificar se há preços vencidos
            preco_vencido = False
            for insumo_id in insumos_selecionados:
                insumo = Insumo.objects.get(id=insumo_id)
                precos_recentes = Preco.objects.filter(insumo=insumo).order_by('-data_cotacao')[:3]
                for preco in precos_recentes:
                    if (datetime.now().date() - preco.data_cotacao).days > 180:
                        preco_vencido = True
                        break
                if preco_vencido:
                    break

            # Se houver preços vencidos, emitir alerta
            if preco_vencido:
                messages.warning(request, "Este mapa de cotação possui preços vencidos!")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Relatório de Cotações'

            headers = [
                'Código', 'Descrição', 'Unidade', 'Total', 'Qtd Cotações', 'Dados', 'Empresa 1', 'Empresa 2',
                'Empresa 3'
            ]

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            row_num = 2
            empresa_columns = {'Empresa 1': 7, 'Empresa 2': 8, 'Empresa 3': 9}

            for insumo_id in insumos_selecionados:
                insumo = Insumo.objects.get(id=insumo_id)
                precos_recentes = Preco.objects.filter(insumo=insumo).order_by('-data_cotacao')[:3]

                if precos_recentes:
                    quantidade_cotacoes = len(precos_recentes)
                    precos = [preco.preco for preco in precos_recentes]
                    calculo = 0
                    if calcular == 'media' and precos:
                        calculo = statistics.mean(precos)
                    elif calcular == 'mediana' and precos:
                        calculo = statistics.median(precos)

                    ws.cell(row=row_num, column=1, value=insumo.codigo).border = thin_border
                    ws.cell(row=row_num, column=2, value=insumo.nome).border = thin_border
                    ws.cell(row=row_num, column=3, value=insumo.unidade_medida).border = thin_border
                    ws.cell(row=row_num, column=4, value=calcular).border = thin_border
                    ws.cell(row=row_num + 1, column=4, value=calculo).border = thin_border
                    ws.cell(row=row_num, column=5, value=len(precos_recentes)).border = thin_border

                    fornecedores = []
                    cnpjs = []
                    datas = []
                    contatos = []
                    valores = []
                    vendedores = []

                    for preco in precos_recentes:
                        fornecedores.append(preco.razao_social)
                        cnpjs.append(preco.cnpj)
                        datas.append(preco.data_cotacao)
                        contatos.append(preco.telefone)
                        valores.append(preco.preco)
                        vendedores.append(preco.vendedor)

                    for idx, (fornecedor, cnpj, data, contato, valor, vendedor) in enumerate(
                            zip(fornecedores, cnpjs, datas, contatos, valores, vendedores)):
                        if idx < len(empresa_columns):
                            empresa, empresa_column = list(empresa_columns.items())[idx]
                            ws.cell(row=row_num, column=empresa_column, value=fornecedor).border = thin_border
                            ws.cell(row=row_num + 1, column=empresa_column, value=cnpj).border = thin_border

                            if (datetime.now().date() - data).days > 180:
                                ws.cell(row=row_num + 2, column=empresa_column, value=data).border = thin_border
                                ws.cell(row=row_num + 2, column=empresa_column).fill = openpyxl.styles.PatternFill(
                                    start_color='FF0000',
                                    end_color='FF0000',
                                    fill_type='solid'
                                )
                            else:
                                ws.cell(row=row_num + 2, column=empresa_column, value=data).border = thin_border

                            ws.cell(row=row_num + 3, column=empresa_column, value=contato).border = thin_border
                            ws.cell(row=row_num + 4, column=empresa_column, value=vendedor).border = thin_border
                            ws.cell(row=row_num + 5, column=empresa_column, value=valor).border = thin_border

                    ws.cell(row=row_num, column=6, value="Fornecedor:").border = thin_border
                    ws.cell(row=row_num + 1, column=6, value="CNPJ:").border = thin_border
                    ws.cell(row=row_num + 2, column=6, value="Data:").border = thin_border
                    ws.cell(row=row_num + 3, column=6, value="Contato:").border = thin_border
                    ws.cell(row=row_num + 4, column=6, value="Vendedor:").border = thin_border
                    ws.cell(row=row_num + 5, column=6, value="Valor:").border = thin_border

                    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num + 5, end_column=1)
                    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num + 5, end_column=2)
                    ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num + 5, end_column=3)
                    ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num + 4, end_column=4)
                    ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num + 5, end_column=5)

                    for i in range(4):
                        cell = ws.cell(row=row_num + i, column=4)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    ws.cell(row=row_num + 5, column=4, value=calculo).border = thin_border
                    ws.cell(row=row_num + 5, column=4).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_num + 5, column=4).number_format = 'R$ #,##0.00'

                    for col_num in range(1, 6):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    row_num += 7

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            ws.cell(row=row_num - 1, column=4).font = Font(bold=True)

            temp_dir = tempfile.mkdtemp()

            # Salvando o arquivo Excel temporariamente
            timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
            excel_filename = f'relatorio_cotacoes_{timestamp}.xlsx'
            excel_path = os.path.join(temp_dir, excel_filename)
            wb.save(excel_path)

            # Criar Mapa com o arquivo Excel
            mapa = Mapa.objects.create(usuario=request.user, empresa=request.user.empresa)
            mapa.arquivo.save(excel_filename, File(open(excel_path, 'rb')))

            # Criar diretório para PDFs
            pdf_dir = os.path.join(temp_dir, 'pdfs')
            os.makedirs(pdf_dir, exist_ok=True)

            # Coletar arquivos anexados e ajustar o nome da pasta
            for insumo_id in insumos_selecionados:
                insumo = Insumo.objects.get(id=insumo_id)
                precos = Preco.objects.filter(insumo=insumo).order_by('-data_cotacao')[:3]

                for preco in precos:
                    arquivos = ArquivoAnexado.objects.filter(preco=preco)

                    for arquivo in arquivos:
                        nome_pasta = f"{insumo.codigo}-{insumo.nome.replace(' ', '_')}"
                        insumo_pdf_dir = os.path.join(pdf_dir, nome_pasta)
                        os.makedirs(insumo_pdf_dir, exist_ok=True)

                        # Copiar o arquivo para o diretório
                        shutil.copy(arquivo.arquivo.path, insumo_pdf_dir)

            # Criando o arquivo ZIP com os PDFs e o arquivo Excel
            zip_file_path = os.path.join(temp_dir, 'arquivos.zip')
            with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                # Adicionar o arquivo Excel
                zipf.write(excel_path, os.path.basename(excel_path))

                # Adicionar os PDFs
                for root, dirs, files in os.walk(pdf_dir):
                    for file in files:
                        zipf.write(os.path.join(root, file), os.path.relpath(
                            os.path.join(root, file), os.path.join(pdf_dir, '..')))

            # Limpando arquivos temporários
            shutil.rmtree(pdf_dir)

            # Enviando o arquivo ZIP como resposta HTTP
            with open(zip_file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename={os.path.basename(zip_file_path)}'

            shutil.rmtree(temp_dir)
            return response

        else:
            # Quando for uma requisição GET
            pesquisa_tipo = request.POST.get('pesquisa_tipo')
            pesquisa_termo = request.POST.get('pesquisa_termo', '').strip().lower()
            insumos_do_usuario = Insumo.objects.filter(empresa=request.user.empresa)

            if pesquisa_tipo == 'descricao':
                insumos_do_usuario = insumos_do_usuario.filter(nome__icontains=pesquisa_termo)
            elif pesquisa_tipo == 'codigo':
                insumos_do_usuario = insumos_do_usuario.filter(codigo__icontains=pesquisa_termo)

            context = {'insumos_do_usuario': insumos_do_usuario}

            return render(request, 'relatorios.html', context)  # Retorna a resposta correta para o GET

    else:
        return render(request, 'error.html', {
            'error_message': 'Você não possui uma empresa associada. Por favor, entre em contato com o administrador.'
        })

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Orcamento, Pasta, ComposicaoSinapi, ComposicaoSicro3, Insumo
from django.db.models import Q

@login_required
def orcamentos_view(request):
    # Lista de orçamentos do usuário
    orcamentos = Orcamento.objects.filter(usuario=request.user)
    pastas = Pasta.objects.filter(usuario=request.user)
    return render(request, 'orcamentos.html', {
        'orcamentos': orcamentos,
        'pastas': pastas
    })


@login_required
def criar_orcamento(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        descricao = request.POST.get('descricao')

        orcamento = Orcamento.objects.create(
            usuario=request.user,
            nome=nome,
            descricao=descricao
        )

        return redirect('orcamentos')

    return render(request, 'criar_orcamento.html')


@login_required
def criar_pasta(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        descricao = request.POST.get('descricao')

        pasta = Pasta.objects.create(
            usuario=request.user,
            nome=nome,
            descricao=descricao
        )

        return redirect('orcamentos')

    return render(request, 'criar_pasta.html')


@login_required
def lista_composicoes(request):
    query = request.GET.get('q', '')
    sinapi_composicoes = ComposicaoSinapi.objects.filter(descricao__icontains=query) if query else ComposicaoSinapi.objects.all()
    sicro3_composicoes = ComposicaoSicro3.objects.filter(descricao__icontains=query) if query else ComposicaoSicro3.objects.all()

    return render(request, 'lista_composicoes.html', {
        'sinapi_composicoes': sinapi_composicoes,
        'sicro3_composicoes': sicro3_composicoes,
        'query': query
    })

@login_required
def lista_insumos(request):
    query = request.GET.get('q', '')
    insumos = Insumo.objects.filter(Q(codigo__icontains=query) | Q(descricao__icontains=query)) if query else Insumo.objects.all()

    return render(request, 'lista_insumos.html', {
        'insumos': insumos,
        'query': query
    })

@login_required
def orcamento_lista(request):
    # Obter os orçamentos e pastas relacionados ao usuário logado
    orcamentos = Orcamento.objects.filter(usuario=request.user)
    pastas = Pasta.objects.filter(usuario=request.user)

    # Passar esses objetos para o template
    context = {
        'orcamentos': orcamentos,
        'pastas': pastas
    }

    return render(request, 'orcamento_lista.html', context)

def success(request):
    return render(request, 'success.html')

from .forms import ProfilePictureForm


from django.shortcuts import render, redirect
from .forms import ProfileForm
from django.contrib.auth.decorators import login_required


@login_required
def update_photo(request):
    user = request.user  # ou use CustomUser se for um modelo personalizado
    if request.method == 'POST' and request.FILES.get('photo'):
        form = ProfileForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            return redirect('profile')  # Redireciona após salvar
    else:
        form = ProfileForm(instance=user)
    return render(request, 'profile_update.html', {'form': form})

@login_required
def change_profile_picture(request):
    if request.method == 'POST' and request.FILES['photo']:
        form = ProfilePictureForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect('profile')  # Redireciona para a página do perfil ou outra página
    else:
        form = ProfilePictureForm(instance=request.user)
    return render(request, 'profile/change_picture.html', {'form': form})